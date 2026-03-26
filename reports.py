# reports.py
# 报表生成模块 - 建筑工程造价成本分析报表输出
# Author: 黑莓 🫐

import os
import math
from typing import List, Dict, Any, Optional
from datetime import datetime
from dataclasses import dataclass, asdict

# Excel 支持
try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    OPENXLSX_AVAILABLE = True
except ImportError:
    OPENXLSX_AVAILABLE = False

from data_generator import (
    ProjectData, MonthlyTrendData, COST_TYPES, PROJECT_TYPES,
)
from analytics import (
    CostOverrunDetector, TrendAnalyzer, CostDistributionAnalyzer,
    ExecutionRateAnalyzer, CostPredictor, AnomalyDetector,
    ComprehensiveAnalyzer, TrendResult, CostAlert, DistributionResult,
    PredictionResult, AnomalyResult,
)


# ============================================================
# 样式配置
# ============================================================

@dataclass
class ReportStyle:
    """报表样式配置"""
    title_font_size: int = 14
    header_font_size: int = 11
    data_font_size: int = 10

    title_font_color: str = "FFFFFF"
    header_font_color: str = "FFFFFF"
    data_font_color: str = "000000"

    title_fill_color: str = "2E5090"    # 深蓝色
    header_fill_color: str = "4472C4"   # 蓝色
    alt_row_color: str = "DCE6F1"       # 浅蓝灰

    # 超支/节省颜色
    overrun_fill: str = "FF6B6B"        # 红色
    savings_fill: str = "69DB7C"        # 绿色
    normal_fill: str = "FFFFFF"         # 白色

    border_style: str = "thin"
    border_color: str = "AAAAAA"

    # 预警级别颜色
    alert_normal: str = "92D050"   # 绿色
    alert_attention: str = "FFFF00"  # 黄色
    alert_warning: str = "FFC000"   # 橙色
    alert_severe: str = "FF0000"    # 红色


# ============================================================
# 样式工具函数
# ============================================================

def _get_style() -> ReportStyle:
    return ReportStyle()


def _apply_border(cell, style: ReportStyle):
    """给单元格添加边框"""
    side = Side(style=style.border_style, color=style.border_color)
    cell.border = Border(left=side, right=side, top=side, bottom=side)


def _style_header_cell(cell, text: str, style: ReportStyle):
    """样式化表头单元格"""
    cell.value = text
    cell.font = Font(bold=True, size=style.header_font_size, color=style.header_font_color)
    cell.fill = PatternFill("solid", fgColor=style.header_fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _apply_border(cell, style)


def _style_title_cell(cell, text: str, style: ReportStyle, cols_span: int = 1):
    """样式化标题单元格"""
    cell.value = text
    cell.font = Font(bold=True, size=style.title_font_size, color=style.title_font_color)
    cell.fill = PatternFill("solid", fgColor=style.title_fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    _apply_border(cell, style)


def _style_data_cell(cell, value: Any, style: ReportStyle, alt_row: bool = False):
    """样式化数据单元格"""
    cell.value = value
    cell.font = Font(size=style.data_font_size, color=style.data_font_color)
    if alt_row:
        cell.fill = PatternFill("solid", fgColor=style.alt_row_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    _apply_border(cell, style)


def _money(value: float) -> str:
    """格式化金额（万元）"""
    return f"{value / 10000:,.2f} 万"


def _percent(value: float) -> str:
    """格式化百分比"""
    return f"{value:+.2f}%"


def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    """生成安全的 Excel 工作表名称"""
    # Excel sheet name cannot contain: : \ / * ? [ ]
    safe = name.replace(":", "-").replace("\\", "-").replace("/", "-")
    safe = safe.replace("*", "-").replace("?", "-").replace("[", "(").replace("]", ")")
    return safe[:max_len]


# ============================================================
# 基础报表类
# ============================================================

class BaseReport:
    """报表基类"""

    def __init__(self, style: Optional[ReportStyle] = None):
        self.style = style or _get_style()
        self.wb = openpyxl.Workbook()
        self._created_at = datetime.now().strftime("%Y-%m-%d %H:%M")

    def _add_metadata(self, ws, title: str, ncols: int):
        """添加报表元信息（标题、时间）"""
        # 标题行
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        title_cell = ws.cell(row=1, column=1)
        _style_title_cell(title_cell, title, self.style)

        # 时间行
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        time_cell = ws.cell(row=2, column=1)
        time_cell.value = f"生成时间: {self._created_at}"
        time_cell.font = Font(size=9, italic=True, color="666666")
        time_cell.alignment = Alignment(horizontal="right")

        return 3  # 数据起始行

    def _auto_column_width(self, ws, min_width: int = 8, max_width: int = 40):
        """自动调整列宽"""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    cell_len = len(str(cell.value)) if cell.value else 0
                    max_len = max(max_len, cell_len)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_width, max(min_width, max_len + 2))


# ============================================================
# 成本汇总报表
# ============================================================

class CostSummaryReport(BaseReport):
    """成本汇总报表"""

    def generate(
        self,
        projects: List[ProjectData],
        filename: Optional[str] = None,
    ) -> str:
        """
        生成成本汇总报表

        Args:
            projects: 项目数据列表
            filename: 保存路径

        Returns:
            保存路径
        """
        ws = self.wb.active
        ws.title = _safe_sheet_name("成本汇总")

        # === 汇总概览 ===
        total_budget = sum(p.预算金额 for p in projects)
        total_actual = sum(p.实际成本 for p in projects)
        total_overrun = total_actual - total_budget
        overrun_rate = (total_overrun / total_budget * 100) if total_budget > 0 else 0
        overall_rate = ExecutionRateAnalyzer.compute_overall_execution_rate(projects)

        # === 概览数据 ===
        start_row = self._add_metadata(ws, "📊 建筑工程造价成本汇总报表", ncols=4)

        overview_labels = ["总预算", "总实际成本", "超支金额", "超支比例", "总体执行率"]
        overview_values = [
            _money(total_budget),
            _money(total_actual),
            _money(total_overrun),
            _percent(overrun_rate),
            f"{overall_rate:.2f}%",
        ]

        for i, (label, value) in enumerate(zip(overview_labels, overview_values)):
            row = start_row + i
            lc = ws.cell(row=row, column=1, value=label)
            lc.font = Font(bold=True, size=self.style.data_font_size)
            lc.fill = PatternFill("solid", fgColor=self.style.alt_row_color)
            _apply_border(lc, self.style)

            vc = ws.cell(row=row, column=2, value=value)
            vc.fill = PatternFill("solid", fgColor=self.style.alt_row_color)
            _apply_border(vc, self.style)
            if "超支" in label:
                vc.font = Font(size=self.style.data_font_size, color="FF0000" if total_overrun > 0 else "008000")

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 20

        # === 明细表格 ===
        detail_start = start_row + len(overview_labels) + 2
        ws.merge_cells(start_row=detail_start - 1, start_column=1, end_row=detail_start - 1, end_column=6)
        title_c = ws.cell(row=detail_start - 1, column=1, value="📋 成本明细")
        _style_title_cell(title_c, "📋 成本明细", self.style)

        headers = ["项目编号", "项目名称", "成本类型", "预算(万元)", "实际成本(万元)", "超支比例"]
        for col, h in enumerate(headers, 1):
            _style_header_cell(ws.cell(row=detail_start, column=col), h, self.style)

        # 按超支金额降序
        sorted_projects = sorted(projects, key=lambda p: p.超支金额, reverse=True)
        for idx, p in enumerate(sorted_projects):
            row = detail_start + 1 + idx
            alt = idx % 2 == 1
            _style_data_cell(ws.cell(row=row, column=1), p.项目编号, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=2), p.项目名称, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=3), p.成本类型, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=4), p.预算金额 / 10000, self.style, alt)
            c = ws.cell(row=row, column=5, value=p.实际成本 / 10000)
            _style_data_cell(c, p.实际成本 / 10000, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=6), _percent(p.超支比例), self.style, alt)

            # 超支比例着色
            rate_cell = ws.cell(row=row, column=6)
            if p.超支比例 > 20:
                rate_cell.fill = PatternFill("solid", fgColor=self.style.overrun_fill)
                rate_cell.font = Font(bold=True, color="FFFFFF")
            elif p.超支比例 < -5:
                rate_cell.fill = PatternFill("solid", fgColor=self.style.savings_fill)

        # 列宽
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 14

        filepath = filename or os.path.join(
            os.path.dirname(__file__), "reports",
            f"成本汇总_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
        os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
        self.wb.save(filepath)
        return filepath


# ============================================================
# 项目对比报表
# ============================================================

class ProjectComparisonReport(BaseReport):
    """项目对比报表"""

    def generate(
        self,
        projects: List[ProjectData],
        filename: Optional[str] = None,
    ) -> str:
        """
        生成项目对比报表

        对比维度：项目间执行率、超支排名、成本类型占比
        """
        ws = self.wb.active
        ws.title = "项目对比"

        # 按项目汇总
        project_summary = CostOverrunDetector.summarize_by_project(projects)

        # 取TOP10超支和TOP10最节省
        sorted_projects = sorted(
            project_summary.values(),
            key=lambda x: x["超支比例"],
            reverse=True,
        )
        worst = sorted_projects[:10]
        best = sorted_projects[-10:][::-1]  # 升序后反转（最节省在前）

        ncols = 7
        start_row = self._add_metadata(ws, "🔍 工程项目成本对比分析报表", ncols=ncols)

        # === TOP10 超支榜 ===
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=ncols)
        _style_title_cell(ws.cell(row=start_row, column=1), "⚠️ 超支TOP10", self.style)

        headers = ["排名", "项目名称", "项目类型", "城市", "总预算(万)", "总实际(万)", "超支比例"]
        header_row = start_row + 1
        for col, h in enumerate(headers, 1):
            _style_header_cell(ws.cell(row=header_row, column=col), h, self.style)

        for idx, p in enumerate(worst):
            row = header_row + 1 + idx
            alt = idx % 2 == 1
            _style_data_cell(ws.cell(row=row, column=1), idx + 1, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=2), p["项目名称"], self.style, alt)
            _style_data_cell(ws.cell(row=row, column=3), p["项目类型"], self.style, alt)
            _style_data_cell(ws.cell(row=row, column=4), p["城市"], self.style, alt)
            _style_data_cell(ws.cell(row=row, column=5), round(p["总预算"] / 10000, 2), self.style, alt)
            _style_data_cell(ws.cell(row=row, column=6), round(p["总实际成本"] / 10000, 2), self.style, alt)
            rc = ws.cell(row=row, column=7, value=_percent(p["超支比例"]))
            _style_data_cell(rc, _percent(p["超支比例"]), self.style, alt)
            rc.fill = PatternFill("solid", fgColor=self.style.overrun_fill)
            rc.font = Font(bold=True, color="FFFFFF")

        # === TOP10 节省榜 ===
        saving_start = header_row + len(worst) + 3
        ws.merge_cells(start_row=saving_start, start_column=1, end_row=saving_start, end_column=ncols)
        _style_title_cell(ws.cell(row=saving_start, column=1), "✅ 节省TOP10", self.style)

        header_row2 = saving_start + 1
        for col, h in enumerate(headers, 1):
            _style_header_cell(ws.cell(row=header_row2, column=col), h, self.style)

        for idx, p in enumerate(best):
            row = header_row2 + 1 + idx
            alt = idx % 2 == 1
            _style_data_cell(ws.cell(row=row, column=1), idx + 1, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=2), p["项目名称"], self.style, alt)
            _style_data_cell(ws.cell(row=row, column=3), p["项目类型"], self.style, alt)
            _style_data_cell(ws.cell(row=row, column=4), p["城市"], self.style, alt)
            _style_data_cell(ws.cell(row=row, column=5), round(p["总预算"] / 10000, 2), self.style, alt)
            _style_data_cell(ws.cell(row=row, column=6), round(p["总实际成本"] / 10000, 2), self.style, alt)
            rc = ws.cell(row=row, column=7, value=_percent(p["超支比例"]))
            _style_data_cell(rc, _percent(p["超支比例"]), self.style, alt)
            rc.fill = PatternFill("solid", fgColor=self.style.savings_fill)

        # === 执行率排名 ===
        rate_start = header_row2 + len(best) + 3
        ws.merge_cells(start_row=rate_start, start_column=1, end_row=rate_start, end_column=ncols)
        _style_title_cell(ws.cell(row=rate_start, column=1), "📊 执行率完整排名", self.style)

        ranked = ExecutionRateAnalyzer.rank_projects_by_execution_rate(projects, top_n=20)
        header_row3 = rate_start + 1
        rank_headers = ["排名", "项目名称", "项目类型", "总预算(万)", "总实际(万)", "执行率", "状态"]
        for col, h in enumerate(rank_headers, 1):
            _style_header_cell(ws.cell(row=header_row3, column=col), h, self.style)

        for idx, r in enumerate(ranked):
            row = header_row3 + 1 + idx
            alt = idx % 2 == 1
            _style_data_cell(ws.cell(row=row, column=1), idx + 1, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=2), r["项目名称"], self.style, alt)
            _style_data_cell(ws.cell(row=row, column=3), r["项目类型"], self.style, alt)
            _style_data_cell(ws.cell(row=row, column=4), round(r["总预算"] / 10000, 2), self.style, alt)
            _style_data_cell(ws.cell(row=row, column=5), round(r["总实际成本"] / 10000, 2), self.style, alt)
            _style_data_cell(ws.cell(row=row, column=6), f"{r['执行率']:.2f}%", self.style, alt)
            status = "超支" if r["执行率"] > 100 else "节省"
            status_cell = ws.cell(row=row, column=7, value=status)
            _style_data_cell(status_cell, status, self.style, alt)
            status_cell.fill = PatternFill(
                "solid",
                fgColor=self.style.overrun_fill if r["执行率"] > 100 else self.style.savings_fill
            )
            status_cell.font = Font(bold=True, color="FFFFFF")

        # 列宽
        for col, w in zip(range(1, ncols + 1), [6, 22, 12, 10, 14, 14, 10]):
            ws.column_dimensions[get_column_letter(col)].width = w

        filepath = filename or os.path.join(
            os.path.dirname(__file__), "reports",
            f"项目对比_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
        os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
        self.wb.save(filepath)
        return filepath


# ============================================================
# 趋势分析报表
# ============================================================

class TrendAnalysisReport(BaseReport):
    """趋势分析报表"""

    def generate(
        self,
        trends: List[MonthlyTrendData],
        projects: Optional[List[ProjectData]] = None,
        filename: Optional[str] = None,
    ) -> str:
        """
        生成趋势分析报表

        Args:
            trends: 月度趋势数据
            projects: 可选，项目数据（用于补充分析）
            filename: 保存路径
        """
        ws = self.wb.active
        ws.title = "趋势分析"

        ncols = 6
        start_row = self._add_metadata(ws, "📈 造价趋势分析报表", ncols=ncols)

        # === 月度明细 ===
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=ncols)
        _style_title_cell(ws.cell(row=start_row, column=1), "月度成本明细", self.style)

        headers = ["月份", "项目类型", "总预算(万元)", "总实际成本(万元)", "执行率", "趋势"]
        header_row = start_row + 1
        for col, h in enumerate(headers, 1):
            _style_header_cell(ws.cell(row=header_row, column=col), h, self.style)

        ta = TrendAnalyzer()
        sorted_trends = sorted(trends, key=lambda t: t.年月)

        # 计算简单趋势（与上期比）
        trend_labels = []
        for i, t in enumerate(sorted_trends):
            if i == 0:
                trend_labels.append("-")
            else:
                prev = sorted_trends[i - 1]
                if t.执行率 > prev.执行率 + 1:
                    trend_labels.append("↑")
                elif t.执行率 < prev.执行率 - 1:
                    trend_labels.append("↓")
                else:
                    trend_labels.append("→")

        for idx, (t, tl) in enumerate(zip(sorted_trends, trend_labels)):
            row = header_row + 1 + idx
            alt = idx % 2 == 1
            _style_data_cell(ws.cell(row=row, column=1), t.年月, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=2), t.项目类型, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=3), round(t.总预算 / 10000, 2), self.style, alt)
            _style_data_cell(ws.cell(row=row, column=4), round(t.总实际成本 / 10000, 2), self.style, alt)
            rc = ws.cell(row=row, column=5, value=f"{t.执行率:.2f}%")
            _style_data_cell(rc, f"{t.执行率:.2f}%", self.style, alt)
            # 执行率着色
            if t.执行率 > 105:
                rc.fill = PatternFill("solid", fgColor=self.style.overrun_fill)
                rc.font = Font(color="FFFFFF", bold=True)
            elif t.执行率 < 95:
                rc.fill = PatternFill("solid", fgColor=self.style.savings_fill)
            _style_data_cell(ws.cell(row=row, column=6), tl, self.style, alt)

        # === 季度汇总 ===
        quarterly = ta.analyze_quarterly(trends)
        qs_start = header_row + len(trends) + 3
        ws.merge_cells(start_row=qs_start, start_column=1, end_row=qs_start, end_column=ncols)
        _style_title_cell(ws.cell(row=qs_start, column=1), "季度汇总", self.style)

        q_headers = ["季度", "总预算(万元)", "总实际(万元)", "执行率", "环比变化", "趋势"]
        q_header_row = qs_start + 1
        for col, h in enumerate(q_headers, 1):
            _style_header_cell(ws.cell(row=q_header_row, column=col), h, self.style)

        for idx, q in enumerate(quarterly):
            row = q_header_row + 1 + idx
            alt = idx % 2 == 1
            _style_data_cell(ws.cell(row=row, column=1), q.周期, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=2), round(q.总预算 / 10000, 2), self.style, alt)
            _style_data_cell(ws.cell(row=row, column=3), round(q.总实际成本 / 10000, 2), self.style, alt)
            _style_data_cell(ws.cell(row=row, column=4), f"{q.总体执行率:.2f}%", self.style, alt)
            qoc_cell = ws.cell(row=row, column=5, value=_percent(q.环比变化))
            _style_data_cell(qoc_cell, _percent(q.环比变化), self.style, alt)
            _style_data_cell(ws.cell(row=row, column=6), q.趋势判断, self.style, alt)

        # === 成本预测 ===
        if len(trends) >= 3:
            preds = CostPredictor.predict_next_n_months(trends, n_future=3)
            pred_start = q_header_row + len(quarterly) + 3
            ws.merge_cells(start_row=pred_start, start_column=1, end_row=pred_start, end_column=ncols)
            _style_title_cell(ws.cell(row=pred_start, column=1), "🔮 未来3个月成本预测", self.style)

            p_headers = ["月份", "预测成本(万元)", "置信下限(万)", "置信上限(万)", "模型说明", ""]
            p_header_row = pred_start + 1
            for col, h in enumerate(p_headers[:ncols], 1):
                _style_header_cell(ws.cell(row=p_header_row, column=col), h, self.style)

            for idx, p in enumerate(preds):
                row = p_header_row + 1 + idx
                alt = idx % 2 == 1
                _style_data_cell(ws.cell(row=row, column=1), p.未来月份, self.style, alt)
                _style_data_cell(ws.cell(row=row, column=2), round(p.预测总成本 / 10000, 2), self.style, alt)
                _style_data_cell(ws.cell(row=row, column=3), round(p.置信区间_下限 / 10000, 2), self.style, alt)
                _style_data_cell(ws.cell(row=row, column=4), round(p.置信区间_上限 / 10000, 2), self.style, alt)
                _style_data_cell(ws.cell(row=row, column=5), p.模型说明[:20], self.style, alt)

        # 列宽
        for col, w in zip(range(1, ncols + 1), [12, 14, 16, 16, 10, 8]):
            ws.column_dimensions[get_column_letter(col)].width = w

        filepath = filename or os.path.join(
            os.path.dirname(__file__), "reports",
            f"趋势分析_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
        os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
        self.wb.save(filepath)
        return filepath


# ============================================================
# 多维度综合报表（多Sheet）
# ============================================================

class ComprehensiveReport(BaseReport):
    """综合报表 - 包含所有分析维度"""

    def generate(
        self,
        projects: List[ProjectData],
        trends: Optional[List[MonthlyTrendData]] = None,
        filename: Optional[str] = None,
    ) -> str:
        """
        生成综合报表（多Sheet）

        Sheet结构：
        1. 概览 - 关键指标仪表盘
        2. 成本汇总 - 明细数据
        3. 超支预警 - 异常项目清单
        4. 成本分布 - 饼图数据表
        5. 趋势分析 - 月度/季度趋势
        6. 预测 - 未来成本预测
        7. 异常检测 - 异常值清单
        """
        wb = self.wb

        # === Sheet 1: 概览 ===
        self._create_overview_sheet(projects, trends)

        # === Sheet 2: 成本汇总 ===
        summary_report = CostSummaryReport(self.style)
        summary_ws = wb.create_sheet(_safe_sheet_name("成本汇总"))
        self._fill_summary_sheet(summary_ws, projects)

        # === Sheet 3: 超支预警 ===
        alert_ws = wb.create_sheet(_safe_sheet_name("超支预警"))
        self._fill_alert_sheet(alert_ws, projects)

        # === Sheet 4: 成本分布 ===
        dist_ws = wb.create_sheet(_safe_sheet_name("成本分布"))
        self._fill_distribution_sheet(dist_ws, projects)

        # === Sheet 5: 趋势分析 ===
        if trends:
            trend_ws = wb.create_sheet(_safe_sheet_name("趋势分析"))
            self._fill_trend_sheet(trend_ws, trends)

        # === Sheet 6: 预测 ===
        if trends and len(trends) >= 3:
            pred_ws = wb.create_sheet(_safe_sheet_name("成本预测"))
            self._fill_prediction_sheet(pred_ws, trends)

        # === Sheet 7: 异常检测 ===
        anomaly_ws = wb.create_sheet(_safe_sheet_name("异常检测"))
        self._fill_anomaly_sheet(anomaly_ws, projects)

        # 删除默认空Sheet
        default_ws = wb.active
        if not default_ws.title or default_ws.max_row < 3:
            wb.remove(default_ws)

        filepath = filename or os.path.join(
            os.path.dirname(__file__), "reports",
            f"综合报表_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
        os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
        self.wb.save(filepath)
        return filepath

    def _create_overview_sheet(self, projects: List[ProjectData], trends):
        """概览仪表盘"""
        ws = self.wb.active
        ws.title = "📊 概览"

        ncols = 4
        start_row = self._add_metadata(ws, "📊 建筑工程造价分析 - 综合概览", ncols=ncols)

        # 关键指标卡片
        total_budget = sum(p.预算金额 for p in projects)
        total_actual = sum(p.实际成本 for p in projects)
        overrun = total_actual - total_budget
        overrun_rate = overrun / total_budget * 100 if total_budget > 0 else 0
        overall_rate = ExecutionRateAnalyzer.compute_overall_execution_rate(projects)
        alerts = CostOverrunDetector.detect_overruns(projects)

        kpis = [
            ("总预算", f"¥ {total_budget/1e6:.2f}M", "1,1"),
            ("总实际成本", f"¥ {total_actual/1e6:.2f}M", "1,3"),
            ("超支金额", f"¥ {overrun/1e6:.2f}M ({overrun_rate:+.1f}%)", "2,1"),
            ("超支项目数", f"{len(alerts)} 个", "2,3"),
        ]

        for label, value, pos in kpis:
            row, col = map(int, pos.split(","))
            row = start_row + row - 1
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
            c = ws.cell(row=row, column=col, value=f"{label}\n{value}")
            c.font = Font(bold=True, size=11, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="2E5090")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(
                left=Side(style="medium", color="FFFFFF"),
                right=Side(style="medium", color="FFFFFF"),
                top=Side(style="medium", color="FFFFFF"),
                bottom=Side(style="medium", color="FFFFFF"),
            )
            ws.row_dimensions[row].height = 30

        # 执行率分布
        dist_start = start_row + 3
        ws.merge_cells(start_row=dist_start, start_column=1, end_row=dist_start, end_column=ncols)
        _style_title_cell(ws.cell(row=dist_start, column=1), "执行率区间分布", self.style)

        dist = ExecutionRateAnalyzer.execution_rate_distribution(projects)
        dist_headers = ["区间", "项目数"]
        for col, h in enumerate(dist_headers, 1):
            _style_header_cell(ws.cell(row=dist_start + 1, column=col), h, self.style)
            ws.merge_cells(start_row=dist_start + 1, start_column=col + 2,
                           end_row=dist_start + 1, end_column=col + 3)

        for idx, (label, count) in enumerate(dist.items()):
            row = dist_start + 2 + idx
            _style_data_cell(ws.cell(row=row, column=1), label, self.style, idx % 2 == 1)
            _style_data_cell(ws.cell(row=row, column=2), count, self.style, idx % 2 == 1)

        # 成本类型TOP5
        top5_start = dist_start + len(dist) + 3
        ws.merge_cells(start_row=top5_start, start_column=1, end_row=top5_start, end_column=ncols)
        _style_title_cell(ws.cell(row=top5_start, column=1), "成本类型TOP5（按实际金额）", self.style)

        pie = CostDistributionAnalyzer.get_pie_chart_data(projects, group_by="cost_type")
        for idx, (label, val, pct) in enumerate(zip(
            pie["labels"][:5], pie["values"][:5], pie["percentages"][:5]
        )):
            row = top5_start + 1 + idx
            alt = idx % 2 == 1
            _style_data_cell(ws.cell(row=row, column=1), label, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=2), f"¥ {val/1e6:.2f}M", self.style, alt)
            _style_data_cell(ws.cell(row=row, column=3), f"{pct:.1f}%", self.style, alt)

        for col, w in zip(range(1, ncols + 1), [20, 18, 14]):
            ws.column_dimensions[get_column_letter(col)].width = w

    def _fill_summary_sheet(self, ws, projects: List[ProjectData]):
        """填充成本汇总Sheet"""
        ncols = 8
        start_row = self._add_metadata(ws, "📋 全量成本明细", ncols=ncols)

        headers = ["项目编号", "项目名称", "项目类型", "城市", "成本类型", "预算(万元)", "实际(万元)", "超支比例"]
        for col, h in enumerate(headers, 1):
            _style_header_cell(ws.cell(row=start_row, column=col), h, self.style)

        sorted_projects = sorted(projects, key=lambda p: p.超支金额, reverse=True)
        for idx, p in enumerate(sorted_projects):
            row = start_row + 1 + idx
            alt = idx % 2 == 1
            vals = [
                p.项目编号, p.项目名称, p.项目类型, p.城市, p.成本类型,
                round(p.预算金额 / 10000, 2),
                round(p.实际成本 / 10000, 2),
                _percent(p.超支比例),
            ]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=v)
                _style_data_cell(c, v, self.style, alt)
                if col == 8:
                    if p.超支比例 > 20:
                        c.fill = PatternFill("solid", fgColor=self.style.overrun_fill)
                        c.font = Font(color="FFFFFF", bold=True)
                    elif p.超支比例 < -5:
                        c.fill = PatternFill("solid", fgColor=self.style.savings_fill)

        for col, w in zip(range(1, ncols + 1), [14, 22, 10, 10, 12, 12, 12, 12]):
            ws.column_dimensions[get_column_letter(col)].width = w

    def _fill_alert_sheet(self, ws, projects: List[ProjectData]):
        """填充超支预警Sheet"""
        ncols = 7
        alerts = CostOverrunDetector.detect_overruns(projects)

        start_row = self._add_metadata(ws, "⚠️ 超支预警清单", ncols=ncols)

        headers = ["项目编号", "项目名称", "成本类型", "超支金额(万元)", "超支比例", "预警级别", "建议"]
        for col, h in enumerate(headers, 1):
            _style_header_cell(ws.cell(row=start_row, column=col), h, self.style)

        level_suggestions = {
            "严重": "立即介入，深度审计",
            "警告": "重点关注，成本管控",
            "关注": "保持监控，分析原因",
        }

        for idx, a in enumerate(alerts):
            row = start_row + 1 + idx
            alt = idx % 2 == 1
            _style_data_cell(ws.cell(row=row, column=1), a.项目编号, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=2), a.项目名称, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=3), a.成本类型, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=4), round(a.超支金额 / 10000, 2), self.style, alt)
            _style_data_cell(ws.cell(row=row, column=5), _percent(a.超支比例), self.style, alt)

            lc = ws.cell(row=row, column=6, value=a.预警级别)
            _style_data_cell(lc, a.预警级别, self.style, alt)
            color_map = {"严重": "FF0000", "警告": "FFC000", "关注": "FFFF00"}
            lc.fill = PatternFill("solid", fgColor=color_map.get(a.预警级别, "92D050"))
            if a.预警级别 in ("严重", "警告"):
                lc.font = Font(bold=True, color="FFFFFF")

            _style_data_cell(ws.cell(row=row, column=7), level_suggestions.get(a.预警级别, ""), self.style, alt)

        for col, w in zip(range(1, ncols + 1), [14, 22, 12, 14, 12, 10, 22]):
            ws.column_dimensions[get_column_letter(col)].width = w

    def _fill_distribution_sheet(self, ws, projects: List[ProjectData]):
        """填充成本分布Sheet"""
        ncols = 5
        start_row = self._add_metadata(ws, "🥧 成本类型分布分析", ncols=ncols)

        # 按成本类型
        headers = ["成本类型", "实际金额(万元)", "占比", "预算(万元)", "执行率"]
        for col, h in enumerate(headers, 1):
            _style_header_cell(ws.cell(row=start_row, column=col), h, self.style)

        pie = CostDistributionAnalyzer.get_pie_chart_data(projects, group_by="cost_type")
        for idx, (label, val, pct, rate) in enumerate(zip(
            pie["labels"], pie["values"], pie["percentages"], pie["execution_rates"]
        )):
            row = start_row + 1 + idx
            alt = idx % 2 == 1
            _style_data_cell(ws.cell(row=row, column=1), label, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=2), round(val / 10000, 2), self.style, alt)
            _style_data_cell(ws.cell(row=row, column=3), f"{pct:.1f}%", self.style, alt)
            _style_data_cell(ws.cell(row=row, column=4), f"{pct:.1f}%", self.style, alt)
            _style_data_cell(ws.cell(row=row, column=5), f"{rate:.2f}%", self.style, alt)

        # 按项目类型
        pt_start = start_row + len(pie["labels"]) + 3
        ws.merge_cells(start_row=pt_start, start_column=1, end_row=pt_start, end_column=ncols)
        _style_title_cell(ws.cell(row=pt_start, column=1), "按项目类型分布", self.style)

        pt_headers = ["项目类型", "实际金额(万元)", "占比", "预算(万元)", "执行率"]
        for col, h in enumerate(pt_headers, 1):
            _style_header_cell(ws.cell(row=pt_start + 1, column=col), h, self.style)

        pt_pie = CostDistributionAnalyzer.get_pie_chart_data(projects, group_by="project_type")
        for idx, (label, val, pct, rate) in enumerate(zip(
            pt_pie["labels"], pt_pie["values"], pt_pie["percentages"], pt_pie["execution_rates"]
        )):
            row = pt_start + 2 + idx
            alt = idx % 2 == 1
            _style_data_cell(ws.cell(row=row, column=1), label, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=2), round(val / 10000, 2), self.style, alt)
            _style_data_cell(ws.cell(row=row, column=3), f"{pct:.1f}%", self.style, alt)
            _style_data_cell(ws.cell(row=row, column=4), f"{pct:.1f}%", self.style, alt)
            _style_data_cell(ws.cell(row=row, column=5), f"{rate:.2f}%", self.style, alt)

        for col, w in zip(range(1, ncols + 1), [18, 16, 10, 14, 12]):
            ws.column_dimensions[get_column_letter(col)].width = w

    def _fill_trend_sheet(self, ws, trends: List[MonthlyTrendData]):
        """填充趋势分析Sheet"""
        ncols = 6
        start_row = self._add_metadata(ws, "📈 月度/季度趋势分析", ncols=ncols)

        ta = TrendAnalyzer()
        quarterly = ta.analyze_quarterly(trends)

        # 月度明细
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=ncols)
        _style_title_cell(ws.cell(row=start_row, column=1), "月度明细", self.style)

        headers = ["月份", "项目类型", "总预算(万元)", "总实际(万元)", "执行率", "趋势"]
        for col, h in enumerate(headers, 1):
            _style_header_cell(ws.cell(row=start_row + 1, column=col), h, self.style)

        sorted_trends = sorted(trends, key=lambda t: t.年月)
        for idx, t in enumerate(sorted_trends):
            row = start_row + 2 + idx
            alt = idx % 2 == 1
            trend_icon = "↑" if idx > 0 and t.执行率 > sorted_trends[idx-1].执行率 + 1 else \
                         ("↓" if idx > 0 and t.执行率 < sorted_trends[idx-1].执行率 - 1 else "→")
            _style_data_cell(ws.cell(row=row, column=1), t.年月, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=2), t.项目类型, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=3), round(t.总预算 / 10000, 2), self.style, alt)
            _style_data_cell(ws.cell(row=row, column=4), round(t.总实际成本 / 10000, 2), self.style, alt)
            rc = ws.cell(row=row, column=5, value=f"{t.执行率:.2f}%")
            _style_data_cell(rc, f"{t.执行率:.2f}%", self.style, alt)
            if t.执行率 > 105:
                rc.fill = PatternFill("solid", fgColor=self.style.overrun_fill)
                rc.font = Font(color="FFFFFF")
            elif t.执行率 < 95:
                rc.fill = PatternFill("solid", fgColor=self.style.savings_fill)
            _style_data_cell(ws.cell(row=row, column=6), trend_icon, self.style, alt)

        # 季度汇总
        qs_row = start_row + len(sorted_trends) + 3
        ws.merge_cells(start_row=qs_row, start_column=1, end_row=qs_row, end_column=ncols)
        _style_title_cell(ws.cell(row=qs_row, column=1), "季度汇总", self.style)

        for col, h in enumerate(["季度", "总预算(万元)", "总实际(万元)", "执行率", "环比", "趋势"], 1):
            _style_header_cell(ws.cell(row=qs_row + 1, column=col), h, self.style)

        for idx, q in enumerate(quarterly):
            row = qs_row + 2 + idx
            alt = idx % 2 == 1
            _style_data_cell(ws.cell(row=row, column=1), q.周期, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=2), round(q.总预算 / 10000, 2), self.style, alt)
            _style_data_cell(ws.cell(row=row, column=3), round(q.总实际成本 / 10000, 2), self.style, alt)
            _style_data_cell(ws.cell(row=row, column=4), f"{q.总体执行率:.2f}%", self.style, alt)
            _style_data_cell(ws.cell(row=row, column=5), _percent(q.环比变化), self.style, alt)
            _style_data_cell(ws.cell(row=row, column=6), q.趋势判断, self.style, alt)

        for col, w in zip(range(1, ncols + 1), [12, 14, 14, 12, 10, 8]):
            ws.column_dimensions[get_column_letter(col)].width = w

    def _fill_prediction_sheet(self, ws, trends: List[MonthlyTrendData]):
        """填充预测Sheet"""
        ncols = 5
        start_row = self._add_metadata(ws, "🔮 未来成本预测", ncols=ncols)

        preds = CostPredictor.predict_next_n_months(trends, n_future=6)

        for col, h in enumerate(["预测月份", "预测成本(万元)", "置信下限(万)", "置信上限(万)", "说明"], 1):
            _style_header_cell(ws.cell(row=start_row, column=col), h, self.style)

        for idx, p in enumerate(preds):
            row = start_row + 1 + idx
            alt = idx % 2 == 1
            _style_data_cell(ws.cell(row=row, column=1), p.未来月份, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=2), round(p.预测总成本 / 10000, 2), self.style, alt)
            _style_data_cell(ws.cell(row=row, column=3), round(p.置信区间_下限 / 10000, 2), self.style, alt)
            _style_data_cell(ws.cell(row=row, column=4), round(p.置信区间_上限 / 10000, 2), self.style, alt)
            _style_data_cell(ws.cell(row=row, column=5), p.模型说明[:30], self.style, alt)

        for col, w in zip(range(1, ncols + 1), [12, 16, 14, 14, 30]):
            ws.column_dimensions[get_column_letter(col)].width = w

    def _fill_anomaly_sheet(self, ws, projects: List[ProjectData]):
        """填充异常检测Sheet"""
        ncols = 6
        anomalies = AnomalyDetector.comprehensive_anomaly_check(projects)

        all_anomalies = anomalies.get("执行率异常", []) + anomalies.get("单价异常", [])

        start_row = self._add_metadata(ws, "🚨 异常值检测报告", ncols=ncols)

        if not all_anomalies:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=ncols)
            c = ws.cell(row=start_row, column=1, value="✅ 未检测到显著异常")
            c.font = Font(size=12, color="008000")
            c.alignment = Alignment(horizontal="center")
            return

        headers = ["项目编号", "项目名称", "检测指标", "实际值", "期望值", "偏差率"]
        for col, h in enumerate(headers, 1):
            _style_header_cell(ws.cell(row=start_row, column=col), h, self.style)

        for idx, a in enumerate(all_anomalies[:30]):  # 最多30条
            row = start_row + 1 + idx
            alt = idx % 2 == 1
            _style_data_cell(ws.cell(row=row, column=1), a.项目编号, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=2), a.项目名称, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=3), a.检测指标, self.style, alt)
            _style_data_cell(ws.cell(row=row, column=4), round(a.实际值, 2), self.style, alt)
            _style_data_cell(ws.cell(row=row, column=5), round(a.期望值, 2), self.style, alt)

            dc = ws.cell(row=row, column=6, value=_percent(a.偏差率))
            _style_data_cell(dc, _percent(a.偏差率), self.style, alt)
            if "超支" in a.异常类型:
                dc.fill = PatternFill("solid", fgColor=self.style.overrun_fill)
                dc.font = Font(color="FFFFFF", bold=True)
            else:
                dc.fill = PatternFill("solid", fgColor=self.style.savings_fill)

        for col, w in zip(range(1, ncols + 1), [14, 22, 16, 12, 12, 12]):
            ws.column_dimensions[get_column_letter(col)].width = w


# ============================================================
# 文本报表生成器（无需openpyxl）
# ============================================================

class TextReport:
    """纯文本报表（无依赖版本）"""

    @staticmethod
    def summary_text(projects: List[ProjectData]) -> str:
        """生成文本格式的成本汇总"""
        total_budget = sum(p.预算金额 for p in projects)
        total_actual = sum(p.实际成本 for p in projects)
        overrun = total_actual - total_budget
        overrun_rate = overrun / total_budget * 100 if total_budget > 0 else 0

        lines = [
            "=" * 60,
            "         建筑工程造价成本汇总报表",
            "=" * 60,
            f"  总预算:        ¥ {total_budget/1e6:>12.2f} M",
            f"  总实际成本:    ¥ {total_actual/1e6:>12.2f} M",
            f"  超支金额:      ¥ {overrun/1e6:>12.2f} M ({overrun_rate:+.2f}%)",
            f"  总体执行率:    {ExecutionRateAnalyzer.compute_overall_execution_rate(projects):>12.2f}%",
            "=" * 60,
            "",
            f"{'项目名称':<22} {'成本类型':<10} {'预算(万)':>12} {'实际(万)':>12} {'超支比例':>10}",
            "-" * 70,
        ]

        sorted_p = sorted(projects, key=lambda x: x.超支金额, reverse=True)[:20]
        for p in sorted_p:
            lines.append(
                f"  {p.项目名称:<20} {p.成本类型:<10} "
                f"{p.预算金额/10000:>12.2f} {p.实际成本/10000:>12.2f} {p.超支比例:>+9.2f}%"
            )

        return "\n".join(lines)

    @staticmethod
    def trend_text(trends: List[MonthlyTrendData]) -> str:
        """生成文本格式的趋势报表"""
        ta = TrendAnalyzer()
        monthly = ta.analyze_monthly_trend(trends)
        quarterly = ta.analyze_quarterly(trends)

        lines = [
            "=" * 60,
            "         造价趋势分析报表",
            "=" * 60,
            f"  分析周期: {monthly.周期}",
            f"  总体执行率: {monthly.总体执行率}%",
            f"  趋势判断: {monthly.趋势判断}",
            f"  环比变化: {monthly.环比变化:+.2f}%",
            "=" * 60,
            "",
            "【季度汇总】",
            f"{'季度':<10} {'总预算(万)':>14} {'总实际(万)':>14} {'执行率':>10} {'趋势':<6}",
            "-" * 56,
        ]

        for q in quarterly:
            lines.append(
                f"  {q.周期:<8} {q.总预算/10000:>14.2f} {q.总实际成本/10000:>14.2f} "
                f"{q.总体执行率:>9.2f}%  {q.趋势判断}"
            )

        return "\n".join(lines)


# ============================================================
# 测试入口
# ============================================================

if __name__ == "__main__":
    from data_generator import ProjectDataGenerator, TimeSeriesGenerator

    print("=" * 60)
    print("测试报表生成")
    print("=" * 60)

    if not OPENXLSX_AVAILABLE:
        print("⚠️ openpyxl 未安装，仅测试文本报表")
    else:
        print("✅ openpyxl 可用")

    # 生成测试数据
    gen = ProjectDataGenerator(seed=42)
    projects = gen.generate_projects(n_projects=30, n_cost_types_per_project=4)

    ts_gen = TimeSeriesGenerator(seed=42)
    trends = ts_gen.generate_monthly_trend("住宅楼", n_months=12)

    # 文本报表测试
    print("\n【文本汇总报表】")
    print(TextReport.summary_text(projects))

    print("\n【文本趋势报表】")
    print(TextReport.trend_text(trends))

    # Excel 报表测试
    if OPENXLSX_AVAILABLE:
        print("\n【生成 Excel 报表】")

        # 成本汇总
        summary_path = CostSummaryReport().generate(projects)
        print(f"  ✅ 成本汇总: {summary_path}")

        # 项目对比
        compare_path = ProjectComparisonReport().generate(projects)
        print(f"  ✅ 项目对比: {compare_path}")

        # 趋势分析
        trend_path = TrendAnalysisReport().generate(trends, projects)
        print(f"  ✅ 趋势分析: {trend_path}")

        # 综合报表
        comp_path = ComprehensiveReport().generate(projects, trends)
        print(f"  ✅ 综合报表: {comp_path}")
    else:
        print("\nopenpyxl 未安装，跳过 Excel 报表测试")
